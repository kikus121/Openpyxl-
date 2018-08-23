"""
Created By: Kamil Buczkowski
Last Modified: 24/07/2018

This file uses Function.py and Lists.py  to fill in XLS from DPOW
"""

import Functions as f1
import DataStructs 
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter




# Read xlx
wb1 = openpyxl.load_workbook('*****.xlsx')#Name of the excel workbook that contains the information
sheetName1 = 'Deliverables' #Name of the sheet in the excel workbook that  contains the information
readRowStart = 112
sheet = wb1[sheetName1]
# Write xlx
wb2 = openpyxl.load_workbook('*****.xlsx')# Name of the excel work book to write the information too
sheetName2 = 'Assets'# Name of the sheet in excel work book to write the information too
sheetName3 = 'Parameters'
writeRowStart = 6

saveAs="Test_2.xlsx" #The name of the excel workbook



"""
Code Start
"""
print("Program Strart")

startRowRead = readRowStart
startRowWrite = writeRowStart

#findComonUniclass(wb1,wb2,sheetName1,sheetName2,Row,saveAs)
f1.findComonUniclass(wb1,wb2,sheetName1,sheetName3,startRowRead,saveAs)

for i in range(0,(sheet.max_row - readRowStart)):
    
    print("Working on Row {}".format(startRowRead))
    #determineCobie(wb,sheetName,Row)
    dic = f1.determineLib(wb1,sheetName1,startRowRead)

    #def writeToDictionary(dictionary,startRow,wb,sheetName)
    f1.writeToDictionary(dic,startRowRead,wb1,sheetName1)

    #writeFromDictionary(dictionary,startRow,wb,sheetName,saveAs)
    startRowWrite = f1.writeFromDictionary(dic,startRowWrite,wb2,sheetName2,saveAs)

    startRowRead += 1
    

#f1.writeToAssets(wb1,sheetName1,wb2,sheetName2,saveAs)
print("Program End new file created caled : {}".format(saveAs))
