import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

wb = openpyxl.load_workbook('RobSheet.xlsx')
sheet = wb['Deliverables']
print(sheet.max_row)
print(sheet.max_column)


