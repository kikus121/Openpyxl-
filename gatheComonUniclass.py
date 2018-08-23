"""
Created By: Kamil Buczkowski
Last Modified: 23/07/2018

"""
#import Lists
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter




def saveWorkbook(workbook,name):
    if name.endswith('.xlsx'):
        name = name[:-5]
    name = name + '_BS_Rules' + '.xlsx'

    workbook.save(name)
              
"""
#Convets "A" to 1
"""
def letterToNum(letter):
    return column_index_from_string(letter)

"""   
#Convets 1 index to "A"
"""
def numToLetter(num):
    return get_column_letter(num)
"""   
#Reads a column and return a list of the values in this column
"""    
def readColumn(column,startRow,endRow,wb,sheetName):
    list = []
    sheet = wb[sheetName]
    for i in range(startRow,endRow+1):
        cell = sheet.cell(row = i, column = column)
        list.append(cell.value)

    return list
"""
Deerminse cobie type and return proper lib ftom LIsts.py
"""
def findComonUniclass(wb1,wb2,sheetName1,sheetName2,Row,saveAs):
    sheet = wb1[sheetName1]
    cell = sheet.cell(row = Row, column = letterToNum("B") )

    systemStr = ""
    floorStr = ""
    spaceStr = ""
    typeStr = ""
    zoneStr = ""
  
    list = readColumn(2,Row,sheet.max_row,wb1,sheetName1)

    for item in list:
        string = str(item)
        string = string[0]+string[1]


        if string == "EF":
             if floorStr:
                 floorStr =  floorStr + "," +item
             else:
                 floorStr = item
        
        elif string == "Sl":
             if spaceStr:
                 spaceStr =  floorStr + "," +item
             else:
                 floorStr = item
        
        elif string == "Ac":
             if zoneStr:
                 zoneStr =  zoneStr + "," +item
             else:
                 zoneStr = item
        
        elif string == "Pr":
             if typeStr:
                 typeStr =  typeStr + "," +item
             else:
                 typeStr = item
        
        elif string == "Ss":
             if systemStr:
                 systemStr =  systemStr + "," +item
             else:
                 systemStr = item

    sheet = wb2[sheetName2]
    
    column = letterToNum("G")
    Row = 8
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = systemStr
    Row +=1
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = floorStr
    Row +=1
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = spaceStr
    Row +=1
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = zoneStr
    Row +=1
     
    cell = sheet.cell(row = Row, column = column )  
    cell.value = typeStr
    Row +=1
 
    saveWorkbook(wb2,saveAs)

