"""
Created By: Kamil Buczkowski
Last Modified: 23/07/2018

This file contains all the functions() used to fill in XLS from DPOW
"""
import DataStructs
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

#saveAs = "Test_Updated2.xlsx"

               
"""
#Convets "A" to 1
"""
def letterToNum(letter):
    return column_index_from_string(letter)

"""   
#Convets 1 index to "A"
"""
def numToLetter(num):
    return get_column_letter(num)

"""    
#Takes in workbook and saves it to new file called "workbookUpdated.xlsx"
"""
def saveWorkbook(workbook,name):
    if name.endswith('.xlsx'):
        name = name[:-5]
    name = name + '_BS_Rules' + '.xlsx'

    workbook.save(name)
"""   
#Reads a column and return a list of the values in this column
"""    
def readColumn(column,startRow,endRow,wb,sheetName):
    list = []
    sheet = wb[sheetName]
    for i in range(startRow,endRow+1):
        cell = sheet.cell(row = i, column = column)
        list.append(cell.value)

    return list


"""
#Takes a list of values and updated a specific woorkbook
"""
def writeColumn(list,column,startRow,wb,sheetName,saveAs):
    index = 0
    sheet = wb[sheetName]
    endRow = startRow + len(list)
    for i in range(startRow,endRow):
        cell = sheet.cell(row = i, column = column)
        
        if list[index]:
            cell.value = list[index]
    
        
        index += 1

    saveWorkbook(wb,saveAs)
"""
Deerminse cobie type and return proper lib ftom LIsts.py
"""
def determineLib(wb,sheetName,Row):
    sheet = wb[sheetName]
    cell = sheet.cell(row = Row, column = letterToNum("B") )
    string = str(cell.value)
    string = string[0]+string[1]
    if string == "EF":
        return DataStructs.floorDic
    
    elif string == "Sl":
        return DataStructs.spaceDic
    
    elif string == "Ac":
        return DataStructs.zoneDic
    
    elif string == "Pr":
        return DataStructs.typeDic
    
    elif string == "Ss":
        return DataStructs.systemDic

    else:
        return False

"""
Takes a lib, looks at its Cobie def. location and returns a list of as folows
list[startCoulmn,startRow,endColumn,endRow]
"""
def findSizeOfBox(lib):
    startCell = lib["excelSheetLocation"]["CellStart"]
    endCell = lib["excelSheetLocation"]["CellEnd"]
    startRow = ""
    endRow = ""
    startColumn = ""
    endColumn = ""
    list = [1,2,3,4]

    #Find start/end column
    for i in  range(0,2):
        startColumn = startColumn + startCell[i] 
        endColumn = endColumn + endCell[i]

    
    startColumn = letterToNum(startColumn)
    endColumn = letterToNum(endColumn)
    list[0] = startColumn
    list[2] = endColumn

    #find start row
    for i in range(-(len(startCell)-2),0):
        startRow = startRow + startCell[i] 

    startRow = int(startRow)
    list[1] = startRow
     
    for i in range(-(len(endCell)-2),0):
        endRow = endRow + endCell[i]  

    endRow = int(endRow)
    list[3] = endRow

    return list    
"""   
# Takes dictonary and writes to an excel file , return the curent cell
"""
        
def writeFromDictionary(dictionary,startRow,wb,sheetName,saveAs):
    if dictionary:
        
        sheet = wb[sheetName]
        currentRow = startRow


        #Fill in Description
        currentCoulumn = column_index_from_string("C")
        cell = sheet.cell(row = startRow, column = currentCoulumn)
        cell.value = dictionary["ID"]["Description"]

        #Fill in discipline
        currentCoulumn = column_index_from_string("D")
        cell = sheet.cell(row = startRow, column = currentCoulumn)
        cell.value = dictionary["ID"]["Role"]

        #Fill in model
        currentCoulumn = column_index_from_string("F")
        cell = sheet.cell(row = startRow, column = currentCoulumn)
        cell.value = dictionary["ID"]["Model"]

        #Fill in BA type
        currentCoulumn = column_index_from_string("I")
        cell = sheet.cell(row = startRow, column = currentCoulumn)
        cell.value = dictionary["ID"]["Classification"]

        
        
        saveWorkbook(wb,saveAs)

        curentRow = startRow + 1

    else: curentRow = startRow
    
    return curentRow
"""
#Copys from a row to fill a dictionary (systemDic, typeDic ets ) return a filled dic
"""
def writeToDictionary(dictionary,startRow,wb,sheetName):
    if dictionary:
        currentRow = startRow
        currentColumn = 2
        sheet = wb[sheetName]
        model = ""

        # Fill in clasification
        cell = sheet.cell(row = currentRow, column = currentColumn)
        dictionary["ID"]["Classification"] = cell.value

        currentColumn += 1
        
        #Fill in description
        cell = sheet.cell(row = currentRow, column = currentColumn)
        dictionary["ID"]["Description"] = cell.value

        #Fill in role
        currentColumn = column_index_from_string("I")
        cell = sheet.cell(row = currentRow, column = currentColumn)
        dictionary["ID"]["Role"] = cell.value

        #Fill in model
        currentColumn = column_index_from_string("D")
        cell = sheet.cell(row = currentRow, column = currentColumn)
       
        if cell.value:

            dictionary["ID"]["Model"] = cell.value
            
       
        for i in range(0,8):
            currentColumn += 1
            cell = sheet.cell(row = currentRow, column = currentColumn)
            if dictionary["ID"]["Model"]:
                if cell.value:
                    dictionary["ID"]["Model"] = dictionary["ID"]["Model"]+"-"+cell.value
                    
            else:dictionary["ID"]["Model"] = cell.value 
    


        #Fill in parameters
        currentColumn = column_index_from_string("CM")
        cell = sheet.cell(row = currentRow, column = currentColumn)
        # v is value. k is parameter
        for k, v in dictionary["Parameters"].items():
            if cell.value :
                dictionary["Parameters"][k]["Stage"] = cell.value
            currentColumn += 1
            cell = sheet.cell(row = currentRow, column = currentColumn)
                

        return dictionary

    else: return False
"""
# Uses DPOW/Deliverables to fill  Assets
"""
assetSratingRow = 112
def writeToAssets(wbRead,sheetRead,wbWrite,sheetWrite,saveAs):
    
    endRow = wbRead[sheetRead].max_row

    list = readColumn(2,assetSratingRow,endRow,wbRead,sheetRead)# Read uniclass
    writeColumn(list,9,6,wbWrite,sheetWrite,saveAs)#Write uniclass

    list = readColumn(3,assetSratingRow,endRow,wbRead,sheetRead)# Read description
    writeColumn(list,3,6,wbWrite,sheetWrite,saveAs)#Write description

    list = readColumn(letterToNum("L"),assetSratingRow,endRow,wbRead,sheetRead)# Read ArchModel
    writeColumn(list,letterToNum("D"),6,wbWrite,sheetWrite,saveAs)#Write ArchModel

    list = readColumn(letterToNum("CM"),assetSratingRow,endRow,wbRead,sheetRead)# Read Notes
    writeColumn(list,letterToNum("H"),6,wbWrite,sheetWrite,saveAs)#Write Notes

    #Read model
    column = 4
    list = readColumn(4,assetSratingRow,endRow,wbRead,sheetRead)
    
    for i in range(1,9):
        column += 1 
        listTemp = readColumn(column,assetSratingRow,endRow,wbRead,sheetRead)

        for k in range(0,len(listTemp)):
            if listTemp[k]:
                list[k] = str(list[k])+'-'+str(listTemp[k]) 
                    
    writeColumn(list,letterToNum("F"),6,wbWrite,sheetWrite,saveAs)#Write model

"""
Searches DPOW for uniclasses and puts comon ones into  one string seperaretd by
a "," it then pastes these strings somewhwere
"""
def findComonUniclass(wb1,wb2,sheetName1,sheetName2,Row,saveAs):
    sheet = wb1[sheetName1]
    cell = sheet.cell(row = Row, column = letterToNum("B") )

    systemStr = ""
    floorStr = ""
    spaceStr = ""
    typeStr = ""
    zoneStr = ""
  
    list = readColumn(2,Row,sheet.max_row,wb1,sheetName1)

    for item in list:
        string = str(item)
        string = string[0]+string[1]


        if string == "EF":
             if floorStr:
                 floorStr =  floorStr + "," +item
             else:
                 floorStr = item
        
        elif string == "Sl":
             if spaceStr:
                 spaceStr =  floorStr + "," +item
             else:
                 floorStr = item
        
        elif string == "Ac":
             if zoneStr:
                 zoneStr =  zoneStr + "," +item
             else:
                 zoneStr = item
        
        elif string == "Pr":
             if typeStr:
                 typeStr =  typeStr + "," +item
             else:
                 typeStr = item
        
        elif string == "Ss":
             if systemStr:
                 systemStr =  systemStr + "," +item
             else:
                 systemStr = item

    sheet = wb2[sheetName2]
    
    column = letterToNum("G")
    Row = 8
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = systemStr
    Row +=1
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = floorStr
    Row +=1
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = spaceStr
    Row +=1
    
    cell = sheet.cell(row = Row, column = column )  
    cell.value = zoneStr
    Row +=1
     
    cell = sheet.cell(row = Row, column = column )  
    cell.value = typeStr
    Row +=1
 
    saveWorkbook(wb2,saveAs)        
   
