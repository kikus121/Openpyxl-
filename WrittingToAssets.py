import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

wbRead = openpyxl.load_workbook('RobSheet.xlsx')
sheetRead = "Deliverables"
wbWrite = openpyxl.load_workbook('SamSource.xlsx')
sheetWrite = "Assets"
#Reads a column and return a list of the values in this column


#Convets "A" to 1

def letterToNum(letter):
    return column_index_from_string(letter)
    
#Convets 1 index to "A"

def numToLetter(num):
    return get_column_letter(num)    
def readColumn(column,startRow,endRow,wb,sheetName):
    list = []
    sheet = wb[sheetName]
    for i in range(startRow,endRow+1):
        cell = sheet.cell(row = i, column = column)
        list.append(cell.value)

    return list
"""
list = readColumn(2,94,120,wbRead,sheetRead)
print(list)
"""
#Takes a list of values and updated a specific woorkbook

def writeColumn(list,column,startRow,wb,sheetName):
    index = 0
    sheet = wb[sheetName]
    endRow = startRow + len(list)
    for i in range(startRow,endRow):
        cell = sheet.cell(row = i, column = column)
        
        if list[index]:
            cell.value = list[index]
        else:
            cell.value = 'None'
     
        
        index += 1
        
    wb.save('Test_Updated.xlsx')
       
def writeToAssets(wbRead,sheetRead,wbWrite,sheetWrite):

    endRow = wbRead[sheetRead].max_row

    list = readColumn(2,94,endRow,wbRead,sheetRead)# Read uniclass
    writeColumn(list,9,6,wbWrite,sheetWrite)#Write uniclass

    list = readColumn(3,94,endRow,wbRead,sheetRead)# Read description
    writeColumn(list,3,6,wbWrite,sheetWrite)#Write description

    list = readColumn(letterToNum("L"),94,endRow,wbRead,sheetRead)# Read ArchModel
    writeColumn(list,letterToNum("D"),6,wbWrite,sheetWrite)#Write ArchModel

    list = readColumn(letterToNum("CM"),94,endRow,wbRead,sheetRead)# Read Notes
    writeColumn(list,letterToNum("H"),6,wbWrite,sheetWrite)#Write Notes

    #Read model
    column = 4
    list = readColumn(4,94,endRow,wbRead,sheetRead)
    
    for i in range(1,9):
        column += 1 
        listTemp = readColumn(column,94,endRow,wbRead,sheetRead)

        for k in range(0,len(listTemp)):
            list[k] = str(list[k])+str(listTemp[k]) 
            
    writeColumn(list,letterToNum("F"),6,wbWrite,sheetWrite)#Write model
    
        
    




 
    return

writeToAssets(wbRead,sheetRead,wbWrite,sheetWrite)

