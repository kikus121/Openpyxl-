import Functions as fl
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

systemDic ={
    "Varibles": {
        "NumberOfRules": 0,
        "NumberOfParameters": 9
    },
    "ID": {
        "Classification": "",
        "Model": "",
        "Role":"",
        "Description": ""
    },
    "excelSheetLocation": {
        "CellStart":"CM74",
        "CellEnd":"CU74"
    },
    "Rules": {
        "Check Multiple Properties": bool(False) ,
        "Check Multiple Properties - m2": bool(False),
        "Check Multiple Properties - mm": bool(False),
        "Check Multiple Properties for Related Element": bool(False),
        "Check Omniclass Depth": bool(False),
        "Check Property Exists": bool(False),
        "Check Property Value Equals": bool(False),
        "Check Property Value Equals Any of Multiple Values": bool(False),
        "Check Property Value Exists": bool(False),
        "Check Property Value Exists Against Quantity": bool(False),
        "Check Property Value Is In Range": bool(False),
        "Check Property Value is Unique": bool(False),
        "Check Property Value Matches Pattern": bool(False),
        "Ensure Property Missing": bool(False),
        "Ensure Property Value Missing": bool(False),
        "Ensure Property Value Not Equal": bool(False),
        "Check Multiple Properties - 2 args": bool(False)   
    },
    "Parameters": {
        "Name": "",
        "CreatedBy": "",
        "CreatedOn": "",
        "Category": "",
        "ComponentNames": "",
        "Extsystem": "",
        "ExtObject": "",
        "ExtIdentifier": "",
        "Description": ""
    }
}
def saveWorkbook(workbook,name):
    if name.endswith('.xlsx'):
        name = name[:-5]
    name = name + '_BS_Rules' + '.xlsx'

    workbook.save(name)
        
def writeFromDictionary(dictionary,startRow,wb,sheetName,saveAs):
    sheet = wb[sheetName]
    currentRow = startRow


    #Fill in Description
    currentCoulumn = column_index_from_string("C")
    cell = sheet.cell(row = startRow, column = currentCoulumn)
    cell.value = dictionary["ID"]["Description"]

    #Fill in discipline
    currentCoulumn = column_index_from_string("D")
    cell = sheet.cell(row = startRow, column = currentCoulumn)
    cell.value = dictionary["ID"]["Role"]

    #Fill in model
    currentCoulumn = column_index_from_string("F")
    cell = sheet.cell(row = startRow, column = currentCoulumn)
    cell.value = dictionary["ID"]["Model"]

    #Fill in BA type
    currentCoulumn = column_index_from_string("I")
    cell = sheet.cell(row = startRow, column = currentCoulumn)
    cell.value = dictionary["ID"]["Classification"]

    
    
    saveWorkbook(wb,saveAs)
    curentRow = startRow + 1
    return curentRow 
         


