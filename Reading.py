import openpyxl

print('Helow world')

wb = openpyxl.load_workbook('Workbook.xlsx')#Creates a workbook object from an excel file

type(wb)
print(wb)

sheetNames = wb.sheetnames #Retutns a listt of all the sheet names in workbook

print(wb.sheetnames)

print(sheetNames[2])


deliverables = wb['Deliverables'] #Looks for the sheet in workbook by name

cell6 = deliverables['C6']

value = cell6.value #Gets the contents of the cell 

print(value)

for i in range(94,428):
    index = 'C'+str(i)
    cell = deliverables[index]
    print("Value:{}  Row:{}  Coordinates:{}".format(cell.value,cell.row,cell.coordinate))
    print("")
    
    
    
   
