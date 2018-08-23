import Functions as fl
import openpyxl
import DataStructs
from openpyxl.utils import column_index_from_string, get_column_letter




def writeToDictionary(dictionary,startRow,wb,sheetName):
    if dictionary:
        currentRow = startRow
        currentColumn = 2
        sheet = wb[sheetName]
        model = ""

        # Fill in clasification
        cell = sheet.cell(row = currentRow, column = currentColumn)
        dictionary["ID"]["Classification"] = cell.value

        currentColumn += 1
        
        #Fill in description
        cell = sheet.cell(row = currentRow, column = currentColumn)
        dictionary["ID"]["Description"] = cell.value

        #Fill in role
        currentColumn = column_index_from_string("I")
        cell = sheet.cell(row = currentRow, column = currentColumn)
        dictionary["ID"]["Role"] = cell.value

        #Fill in model
        currentColumn = column_index_from_string("D")
        cell = sheet.cell(row = currentRow, column = currentColumn)
       
        if cell.value:

            dictionary["ID"]["Model"] = cell.value
            
       
        for i in range(0,8):
            currentColumn += 1
            cell = sheet.cell(row = currentRow, column = currentColumn)
            if dictionary["ID"]["Model"]:
                if cell.value:
                    dictionary["ID"]["Model"] = dictionary["ID"]["Model"]+"-"+cell.value
                    
            else:dictionary["ID"]["Model"] = cell.value 
    


        #Fill in parameters
        currentColumn = column_index_from_string("CM")
        cell = sheet.cell(row = currentRow, column = currentColumn)
        # v is value. k is parameter
        for k, v in dictionary["Parameters"].items():
            if cell.value :
                dictionary["Parameters"][k]["Stage"] = cell.value
            currentColumn += 1
            cell = sheet.cell(row = currentRow, column = currentColumn)
                

        return dictionary

    else: return False     

        
        
    
   
