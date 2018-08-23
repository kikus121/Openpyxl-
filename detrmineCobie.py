import openpyxl
import Lists
from openpyxl.utils import column_index_from_string, get_column_letter

#Convets "A" to 1

def letterToNum(letter):
    return column_index_from_string(letter)

 
#Convets 1 index to "A"

def numToLetter(num):
    return get_column_letter(num)


def determineCobie(wb,sheetName,Row):
    sheet = wb[sheetName]
    cell = sheet.cell(row = Row, column = letterToNum("A") )
    if cell.value == "Fl":
        return Lists.floorDic
    
    elif cell.value == "Sl":
        return Lists.spaceDic
    
    elif cell.value == "Zo":
        return Lists.zoneDic
    
    elif cell.value == "Pr":
        return Lists.typeDic
    
    elif cell.value == "Ss":
        return Lists.systemDic

    else:
        return False
               
